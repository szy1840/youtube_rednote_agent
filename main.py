#!/usr/bin/env python3
"""
TODO: when no speaking in the video, Videolingo logs:No active speech found in audio. The excel status would be: Error in step '🎙️ Transcribing with Whisper'.
Check log at 2025-07-28 19:01.
"""

import asyncio
import logging
import os
import pty
import select
import subprocess
import sys
import time
import json
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
import requests
import srt
from openpyxl import load_workbook

from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow

from config import settings
from email_helper import EmailHelper
from llm_client import LLMClient
from xiaohongshu_selenium import XiaohongshuSelenium

# Configure logging - output to console only (redirected to cron_log.txt by cron job)
def setup_logging():
    """Setup logging configuration for console output only"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[logging.StreamHandler()]  # Console output only
    )

class OllamaManager:
    """Manages Ollama service startup and health checks"""
    
    @staticmethod
    def is_ollama_running() -> bool:
        """Check if Ollama is running by testing the API endpoint"""
        try:
            response = requests.get("http://localhost:11434", timeout=5)
            logging.info("✅ Ollama is already running")
            return True
        except requests.exceptions.RequestException:
            logging.info("❌ Ollama is not running")
            return False
    
    @staticmethod
    def start_ollama() -> bool:
        """Start Ollama with qwen2.5-coder:32b model"""
        try:
            logging.info("🚀 Starting Ollama with qwen2.5-coder:32b...")
            
            # Start ollama run in background
            process = subprocess.Popen(
                ["ollama", "run", "qwen2.5-coder:32b"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            
            # Wait a bit for startup
            time.sleep(10)
            
            # Check if it's running now
            if OllamaManager.is_ollama_running():
                logging.info("✅ Ollama started successfully")
                return True
            else:
                logging.error("❌ Failed to start Ollama")
                return False
                
        except Exception as e:
            logging.error(f"❌ Error starting Ollama: {e}")
            return False

class YouTubePlaylistProcessor:
    """Handles YouTube playlist operations with OAuth2 authentication"""
    
    # YouTube API scopes - we need write access for playlist modification
    SCOPES = ['https://www.googleapis.com/auth/youtube']
    TOKEN_FILE = 'youtube_token.json'
    
    def _find_credentials_file(self) -> str:
        """Find the Google credentials file in the project directory"""
        import glob
        
        # Look for client_secret_*.json files in the current directory
        pattern = "client_secret_*.json"
        matching_files = glob.glob(pattern)
        
        if not matching_files:
            raise FileNotFoundError(f"❌ No Google credentials file found. Please add a file matching pattern: {pattern}")
        
        if len(matching_files) > 1:
            logging.warning(f"⚠️ Multiple credentials files found: {matching_files}. Using the first one: {matching_files[0]}")
        
        credentials_file = matching_files[0]
        logging.info(f"🔐 Found Google credentials file: {credentials_file}")
        return credentials_file
    
    def __init__(self):
        self.youtube = self._get_authenticated_youtube_service()
        self.playlist_id = settings.youtube_playlist_id
    
    def _get_authenticated_youtube_service(self):
        """Get authenticated YouTube service using OAuth2"""
        creds = None
        
        # Load existing token if available
        if Path(self.TOKEN_FILE).exists():
            try:
                creds = Credentials.from_authorized_user_file(self.TOKEN_FILE, self.SCOPES)
                logging.info("🔐 Loaded existing YouTube OAuth2 credentials")
            except Exception as e:
                logging.warning(f"⚠️ Could not load existing credentials: {e}")
                creds = None
        
        # If there are no valid credentials, get new ones
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    logging.info("🔄 Refreshing YouTube OAuth2 credentials...")
                    creds.refresh(Request())
                    logging.info("✅ YouTube credentials refreshed")
                except Exception as e:
                    logging.warning(f"⚠️ Could not refresh credentials: {e}")
                    creds = None
            
            if not creds:
                credentials_file = self._find_credentials_file()
                
                logging.info("🔓 Starting YouTube OAuth2 authentication flow...")
                flow = InstalledAppFlow.from_client_secrets_file(credentials_file, self.SCOPES)
                creds = flow.run_local_server(port=8080)
                logging.info("✅ YouTube OAuth2 authentication completed")
            
            # Save the credentials for the next run
            with open(self.TOKEN_FILE, 'w') as token:
                token.write(creds.to_json())
                logging.info(f"💾 Saved YouTube credentials to {self.TOKEN_FILE}")
        
        return build('youtube', 'v3', credentials=creds)
    
    def get_playlist_videos(self) -> List[Dict[str, Any]]:
        """Get all videos from the configured playlist"""
        try:
            logging.info(f"📺 Fetching videos from playlist: {self.playlist_id}")
            
            videos = []
            next_page_token = None
            
            while True:
                request = self.youtube.playlistItems().list(
                    part='snippet',
                    playlistId=self.playlist_id,
                    maxResults=50,
                    pageToken=next_page_token
                )
                
                response = request.execute()
                
                for item in response['items']:
                    video_info = {
                        'playlist_item_id': item['id'],
                        'video_id': item['snippet']['resourceId']['videoId'],
                        'title': item['snippet']['title'],
                        'url': f"https://www.youtube.com/watch?v={item['snippet']['resourceId']['videoId']}"
                    }
                    videos.append(video_info)
                
                # TODO: get the language of the video, no guarantee that it's the same as the video language
                if 1==0:
                    video_request = self.youtube.videos().list(
                        part='snippet',
                        id=videos[0]['video_id'] # nFkFP1wOPWQ&list=PL5CEFmSqK6fy-7enI3DmJdvFwybpUPDxY
                    )
                    video_response = video_request.execute()

                    if video_response['items']:
                        video_details = video_response['items'][0]['snippet']
                        language = video_details.get('defaultLanguage')  # e.g., 'en', 'zh', 'ja'
                        audio_language = video_details.get('defaultAudioLanguage')  # Audio track language

                next_page_token = response.get('nextPageToken')
                if not next_page_token:
                    break
            
            logging.info(f"✅ Found {len(videos)} videos in playlist")
            return videos
            
        except HttpError as e:
            logging.error(f"❌ YouTube API error: {e}")
            return []
        except Exception as e:
            logging.error(f"❌ Error fetching playlist videos: {e}")
            return []
    
    def remove_video_from_playlist(self, playlist_item_id: str) -> bool:
        """Remove a video from the playlist"""
        try:
            logging.info(f"🗑️ Removing video from playlist: {playlist_item_id}")
            
            request = self.youtube.playlistItems().delete(id=playlist_item_id)
            request.execute()
            
            logging.info("✅ Video removed from playlist successfully")
            return True
            
        except HttpError as e:
            logging.error(f"❌ YouTube API error removing video: {e}")
            return False
        except Exception as e:
            logging.error(f"❌ Error removing video from playlist: {e}")
            return False

class ExcelHandler:
    """Handles Excel file operations for VideoLingo integration"""
    
    def __init__(self, excel_path: str = None):
        if excel_path is None:
            excel_path = settings.videolingo_excel_path
        self.excel_path = Path(excel_path)
    
    def update_video_url(self, video_url: str) -> bool:
        """Update the video URL in row 2, column 1 of the Excel file"""
        try:
            logging.info(f"📊 Updating Excel file with video URL: {video_url}")
            
            # Ensure the Excel file exists
            if not self.excel_path.exists():
                logging.error(f"❌ Excel file not found: {self.excel_path}")
                return False
            
            # Load the workbook
            workbook = load_workbook(self.excel_path)
            worksheet = workbook.active
            
            # Update row 2, column 1 (B1 in Excel notation, but row 2 col 1 in 1-based indexing)
            worksheet.cell(row=2, column=1, value=video_url)
            
            # Clear the status column (row 2, column 5) for new processing
            worksheet.cell(row=2, column=5, value="")
            
            # Save the workbook
            workbook.save(self.excel_path)
            logging.info("✅ Excel file updated successfully")
            return True
            
        except Exception as e:
            logging.error(f"❌ Error updating Excel file: {e}")
            return False
    
    def check_processing_status(self) -> str:
        """Check the processing status in row 2, column 5"""
        try:
            if not self.excel_path.exists():
                logging.error(f"❌ Excel file not found: {self.excel_path}")
                return ""
            
            workbook = load_workbook(self.excel_path)
            worksheet = workbook.active
            
            # Get value from row 2, column 5
            status = worksheet.cell(row=2, column=5).value
            status_str = str(status).strip() if status else ""
            
            logging.info(f"📊 Current processing status: '{status_str}'")
            return status_str
            
        except Exception as e:
            logging.error(f"❌ Error checking Excel status: {e}")
            return ""

class VideoLingoProcessor:
    """Handles VideoLingo processing and monitoring"""
    
    def __init__(self):
        self.videolingo_path = Path(settings.videolingo_path)
        self.excel_handler = ExcelHandler()
        self.base_output_path = self.videolingo_path / "batch" / "output"
        self.current_video_output_path = None
        self.current_video_title = None
    
    def run_videolingo_processing(self, video_url: str, video_title: str) -> bool:
        """Run VideoLingo processing for a video"""
        try:
            # Store the original video title to find the actual folder later
            self.current_video_title = video_title
            self.current_video_output_path = None  # Will be set after processing
            
            logging.info(f"🎬 Processing video: {video_title}")
            
            # Update Excel file with video URL
            if not self.excel_handler.update_video_url(video_url):
                return False
            
            logging.info("🎬 Starting VideoLingo processing...")
            logging.info(f"📁 Working directory: {self.videolingo_path}")
            
            # Run VideoLingo using current environment (already activated by cron job)
            # Check if we're already in a conda environment
            current_env = os.environ.get('CONDA_DEFAULT_ENV', 'None')
            logging.info(f"🌿 Current conda environment: {current_env}")
            
            if current_env == 'videolingo':
                # Use current activated environment directly
                command = [
                    "bash", "-c",
                    f"cd {self.videolingo_path} && python -m batch.utils.batch_processor"
                ]
                logging.info("✅ Using already activated videolingo environment")
            else:
                # Fallback: activate conda environment
                command = [
                    "bash", "-c",
                    f"cd {self.videolingo_path} && source /opt/anaconda3/etc/profile.d/conda.sh && conda activate videolingo && python -m batch.utils.batch_processor"
                ]
                logging.info("🔄 Activating videolingo environment in subprocess")
            
            logging.info("🔄 Starting VideoLingo batch processor...")
            logging.info(f"Command: {' '.join(command)}")
            
            # Prepare environment for subprocess with full PATH and unbuffered output
            env = os.environ.copy()
            
            # Add conda environment paths to ensure ffmpeg is accessible
            conda_env_path = "/opt/anaconda3/envs/videolingo/bin"
            current_path = env.get('PATH', '')
            if conda_env_path not in current_path:
                env['PATH'] = f"{conda_env_path}:{current_path}"
            
            # Force unbuffered output to capture print statements immediately
            env['PYTHONUNBUFFERED'] = '1'
            env['PYTHONIOENCODING'] = 'utf-8'
            
            logging.info(f"🌐 Subprocess PATH: {env['PATH'][:200]}...")  # Log first 200 chars
            
            # Use pty to simulate a real terminal for better output capture
            
            logging.info("🔄 Starting VideoLingo with enhanced log capture (pty + unbuffered)...")
            
            # Create a pseudo-terminal to capture all output including rich.print
            master_fd, slave_fd = pty.openpty()
            
            process = subprocess.Popen(
                command,
                env=env,
                stdout=slave_fd,
                stderr=slave_fd,  # Merge stderr into stdout
                stdin=slave_fd,
                text=True,
                preexec_fn=os.setsid  # Create new process group
            )
            
            # Close slave fd in parent process
            os.close(slave_fd)
            
            # Capture output in real-time using select
            output_lines = []
            try:
                while True:
                    # Check if process is still running
                    if process.poll() is not None:
                        break
                    
                    # Use select to check for available data
                    ready, _, _ = select.select([master_fd], [], [], 0.1)
                    
                    if ready:
                        try:
                            data = os.read(master_fd, 1024).decode('utf-8', errors='ignore')
                            if data:
                                # Split by lines and process each line
                                lines = data.split('\n')
                                for line in lines:
                                    if line.strip():
                                        output_lines.append(line.strip())
                                        # Log to console (which gets redirected to cron_log.txt)
                                        logging.info(f"[VideoLingo] {line.strip()}")
                        except OSError:
                            # Master fd was closed
                            break
                
                # Wait for process to complete and get return code
                return_code = process.wait()
                
                # Read any remaining output
                try:
                    remaining_data = os.read(master_fd, 1024).decode('utf-8', errors='ignore')
                    if remaining_data:
                        for line in remaining_data.split('\n'):
                            if line.strip():
                                output_lines.append(line.strip())
                                logging.info(f"[VideoLingo] {line.strip()}")
                except OSError:
                    pass
                    
            except Exception as e:
                logging.error(f"❌ Error with pty capture: {e}")
                logging.info("🔄 Falling back to standard subprocess capture...")
                
                # Fallback: Use standard subprocess if pty fails
                try:
                    os.close(master_fd)
                except:
                    pass
                
                # Kill the pty process and start a new one with standard pipes
                try:
                    process.terminate()
                    process.wait(timeout=5)
                except:
                    process.kill()
                
                # Restart with standard subprocess
                process = subprocess.Popen(
                    command,
                    env=env,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=0  # Unbuffered
                )
                
                # Standard readline approach
                while True:
                    output = process.stdout.readline()
                    if output == '' and process.poll() is not None:
                        break
                    if output:
                        line = output.strip()
                        output_lines.append(line)
                        logging.info(f"[VideoLingo] {line}")
                
                return_code = process.wait()
            
            finally:
                # Clean up pty resources
                try:
                    os.close(master_fd)
                except:
                    pass
            #

            
            # Log final results
            logging.info(f"🔄 VideoLingo completed with return code: {return_code}")
            logging.info(f"📋 Total output lines captured: {len(output_lines)}")
            
            # Check if the process was successful
            if return_code == 0:
                logging.info("✅ VideoLingo processing completed successfully")
                return True
            else:
                logging.error(f"❌ VideoLingo processing failed with return code {return_code}")
                # Log last few lines for debugging
                if output_lines:
                    logging.error("📋 Last 10 lines of output:")
                    for line in output_lines[-10:]:
                        logging.error(f"   {line}")
                return False
            
        except Exception as e:
            logging.error(f"❌ Error starting VideoLingo processing: {e}")
            return False
    
    def _find_video_output_folder(self) -> Optional[Path]:
        """Find the actual folder that VideoLingo created for this video"""
        try:
            if not self.base_output_path.exists():
                logging.error(f"❌ Output directory doesn't exist: {self.base_output_path}")
                return None
            
            # Get all subdirectories in the output folder
            subdirs = [d for d in self.base_output_path.iterdir() if d.is_dir()]
            
            if not subdirs:
                logging.error(f"❌ No subdirectories found in {self.base_output_path}")
                return None
            
            if not self.current_video_title:
                logging.error(f"❌ No current video title set")
                return None
            
            # Extract alphanumeric characters only from video title
            def extract_alphanumeric(text: str) -> str:
                return ''.join(c for c in text if c.isalnum()).lower()
            
            shortened_title = extract_alphanumeric(self.current_video_title)
            logging.info(f"🔍 Shortened video title: '{shortened_title}'")
            
            # Compare with all folder names
            for folder in subdirs:
                shortened_folder = extract_alphanumeric(folder.name)
                #logging.info(f"🔍 Comparing with folder: '{folder.name}' → '{shortened_folder}'")
                
                if shortened_title == shortened_folder:
                    #logging.info(f"✅ Found matching folder: {folder}")
                    return folder
            
            # Fallback to most recent folder if no match found
            logging.warning(f"⚠️ No matching folder found, using most recent folder")
            latest_dir = max(subdirs, key=lambda d: d.stat().st_mtime)
            logging.info(f"📁 Using fallback folder: {latest_dir}")
            return latest_dir
            
        except Exception as e:
            logging.error(f"❌ Error finding video output folder: {e}")
            return None
    
    def monitor_processing_completion(self, timeout_minutes: int = 5) -> bool:
        """Monitor Excel file for 'Done' status with 10-minute intervals"""
        try:
            logging.info(f"⏰ Monitoring processing completion (timeout: {timeout_minutes} minutes)")
            
            start_time = time.time()
            check_interval = 600  # 10 minutes in seconds
            
            while True:
                # Check current status
                status = self.excel_handler.check_processing_status()
                
                if status.lower() == "done":
                    logging.info("✅ VideoLingo processing completed successfully!")
                    
                    # Now find the actual output folder that VideoLingo created
                    self.current_video_output_path = self._find_video_output_folder()
                    if self.current_video_output_path:
                        logging.info(f"📁 Located output folder: {self.current_video_output_path}")
                    else:
                        logging.error("❌ Could not locate the output folder")
                        return False
                    
                    return True
                
                # Check timeout
                elapsed_minutes = (time.time() - start_time) / 60
                if elapsed_minutes >= timeout_minutes:
                    logging.error(f"❌ Processing timeout after {timeout_minutes} minutes")
                    return False
                
                # Wait 10 minutes before next check
                logging.info(f"⏳ Status: '{status}' - waiting 10 minutes... (elapsed: {elapsed_minutes:.1f}m)")
                time.sleep(check_interval)
                
        except Exception as e:
            logging.error(f"❌ Error monitoring processing: {e}")
            return False
    
    def extract_chinese_text_from_srt(self) -> str:
        """Extract Chinese text from trans.srt file"""
        try:
            if not self.current_video_output_path:
                logging.error(f"❌ No current video output path set")
                return ""
                
            srt_file = self.current_video_output_path / "trans.srt"
            
            if not srt_file.exists():
                logging.error(f"❌ SRT file not found: {srt_file}")
                return ""
            
            logging.info(f"📝 Extracting text from SRT file: {srt_file}")
            
            with open(srt_file, 'r', encoding='utf-8') as f:
                srt_content = f.read()
            
            # Parse SRT content
            subtitles = list(srt.parse(srt_content))
            
            # Extract all text content
            chinese_text = " ".join([subtitle.content for subtitle in subtitles])
            
            logging.info(f"✅ Extracted {len(chinese_text)} characters of Chinese text")
            return chinese_text
            
        except Exception as e:
            logging.error(f"❌ Error extracting text from SRT: {e}")
            return ""
    
    def get_output_video_path(self) -> Optional[Path]:
        """Get the path to the output video file"""
        try:
            if not self.current_video_output_path:
                logging.error(f"❌ No current video output path set")
                return None
                
            video_file = self.current_video_output_path / "output_sub.mp4"
            
            if video_file.exists():
                logging.info(f"✅ Found output video: {video_file}")
                return video_file
            else:
                logging.error(f"❌ Output video not found: {video_file}")
                return None
                
        except Exception as e:
            logging.error(f"❌ Error getting output video path: {e}")
            return None

class VideoProcessor:
    """Main video processing orchestrator"""
    
    def __init__(self):
        self.youtube_processor = YouTubePlaylistProcessor()
        self.videolingo_processor = VideoLingoProcessor()
        self.llm_client = LLMClient()
        self.email_helper = EmailHelper()
    
    async def process_single_video(self, video_info: Dict[str, Any]) -> bool:
        """Process a single video through the complete workflow"""
        try:
            video_title = video_info['title']
            video_url = video_info['url']
            playlist_item_id = video_info['playlist_item_id']
            
            logging.info(f"🎯 Processing video: {video_title}")
            logging.info(f"🔗 URL: {video_url}")
            
            # Set video title for folder matching
            self.videolingo_processor.current_video_title = video_title
            
            # Step 1: Run VideoLingo processing
            if 1==1: # temporary disable VideoLingo processing
                if not self.videolingo_processor.run_videolingo_processing(video_url, video_title):
                    await self.send_error_notification(f"Failed to start VideoLingo processing", video_info)
                    return False
                # pgrep -f "batch\.utils\.batch_processor" | xargs -r kill 
            # Step 2: Monitor for completion
            if not self.videolingo_processor.monitor_processing_completion():
                await self.send_error_notification(f"VideoLingo processing timeout or failed", video_info)
                return False
            
            # Step 3: Extract Chinese text from SRT
            chinese_text = self.videolingo_processor.extract_chinese_text_from_srt()
            if not chinese_text:
                await self.send_error_notification(f"Failed to extract Chinese text from SRT", video_info)
                return False
            
            # Step 4: Generate title and description with ChatGPT
            logging.info("🤖 Generating title and description with ChatGPT...")
            content_result = await self.llm_client.generate_xiaohongshu_content(
                chinese_text, video_title, video_url
            )
            
            if content_result.error_message:
                await self.send_error_notification(f"ChatGPT content generation failed: {content_result.error_message}", video_info)
                return False
            
            # Step 5: Get output video file
            output_video_path = self.videolingo_processor.get_output_video_path()
            if not output_video_path:
                await self.send_error_notification(f"Output video file not found", video_info)
                return False
            
            # Step 6: Save content to text file for manual posting
            logging.info("💾 Saving generated content to text file...")
            content_file_path = self.save_content_to_file(content_result, video_info, output_video_path)
            if not content_file_path:
                await self.send_error_notification(f"Failed to save content to file", video_info)
                return False
            
            # step 7: post everything to xiaohongshu
            # Run hybrid upload process
            uploader = XiaohongshuSelenium()
            success = uploader.run_upload_process(output_video_path, content_result.title, content_result.description)
            if not success:
                await self.send_error_notification(f"Failed to upload video to Xiaohongshu", video_info)
                return False

            # Step 7: Remove video from playlist (optional - don't fail if this fails)
            logging.info("🗑️ Attempting to remove video from YouTube playlist...")  
            try:
                if self.youtube_processor.remove_video_from_playlist(playlist_item_id):
                    logging.info("✅ Video successfully removed from playlist")
                else:
                    logging.warning("⚠️ Failed to remove video from playlist, but continuing...")
            except Exception as e:
                logging.warning(f"⚠️ Error removing video from playlist: {e}, but continuing...")
            
            # Step 8: Send success notification
            await self.send_success_notification(content_result, video_info, str(output_video_path), str(content_file_path))
            
            logging.info(f"🎉 Successfully processed video: {video_title}")
            return True
            
        except Exception as e:
            logging.error(f"❌ Error processing video: {e}")
            await self.send_error_notification(f"Unexpected error: {str(e)}", video_info)
            return False
    
    def save_content_to_file(self, content_result, video_info: Dict[str, Any], output_video_path: Path) -> Optional[Path]:
        """Save generated title and description to a text file for manual posting"""
        try:
            # Create output directory if it doesn't exist
            output_dir = Path("processed_videos")
            output_dir.mkdir(exist_ok=True)
            
            # Create filename with timestamp and video title
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_title = "".join(c for c in video_info['title'] if c.isalnum() or c in (' ', '-', '_')).rstrip()[:50]
            filename = f"{timestamp}_{safe_title}.txt"
            content_file_path = output_dir / filename
            
            # Prepare content
            content = f"""=== Xiaohongshu Content for Manual Posting ===
                        Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

                        Original Video Title: {video_info['title']}
                        YouTube URL: {video_info['url']}
                        Output Video Path: {output_video_path}

                        === GENERATED TITLE ===
                        {content_result.title}

                        === GENERATED DESCRIPTION ===
                        {content_result.description}

                        === INSTRUCTIONS ===
                        1. Open Xiaohongshu and create a new video post
                        2. Upload the video file from: {output_video_path}
                        3. Copy and paste the title above
                        4. Copy and paste the description above
                        5. Publish the post

                        === END ===
                        """
            
            # Save to file
            with open(content_file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            logging.info(f"✅ Content saved to: {content_file_path}")
            return content_file_path
            
        except Exception as e:
            logging.error(f"❌ Error saving content to file: {e}")
            return None
    
    async def send_success_notification(self, content_result, video_info: Dict[str, Any], video_path: str, content_file_path: str):
        """Send success email notification"""
        try:
            # Include content file path in the notification
            enhanced_description = f"""Generated Content Saved To: {content_file_path}

                                    === GENERATED TITLE ===
                                    {content_result.title}

                                    === GENERATED DESCRIPTION ===
                                    {content_result.description}

                                    === MANUAL POSTING INSTRUCTIONS ===
                                    1. Open the content file at: {content_file_path}
                                    2. Follow the instructions in the file to manually post to Xiaohongshu
                                    3. The processed video is available at: {video_path}
                                    """
            
            self.email_helper.send_video_notification(
                chinese_title=f"[READY FOR MANUAL POSTING] {content_result.title}",
                chinese_description=enhanced_description,
                video_path=video_path,
                original_video_title=video_info['title'],
                youtube_url=video_info['url']
            )
            logging.info("✅ Success notification sent")
        except Exception as e:
            logging.error(f"❌ Failed to send success notification: {e}")
    
    async def send_error_notification(self, error_message: str, video_info: Dict[str, Any]):
        """Send error email notification"""
        try:
            self.email_helper.send_error_notification(
                error_message=error_message,
                video_info=video_info,
                original_video_title=video_info['title']
            )
            logging.info("📧 Error notification sent")
        except Exception as e:
            logging.error(f"❌ Failed to send error notification: {e}")
    
    async def process_playlist(self):
        """Process all videos in the YouTube playlist"""
        try:
            logging.info("🚀 Starting YouTube playlist processing...")
            
            # Get all videos from playlist
            videos = self.youtube_processor.get_playlist_videos()
            videos=videos[:1] # Only process the first video
            if not videos:
                logging.error("❌ No videos found in playlist")
                return
            
            logging.info(f"📺 Found {len(videos)} videos to process")
            
            # Check for existing batch processor on Mac
            if sys.platform == "darwin":
                result = subprocess.run(['pgrep', '-f', 'batch.utils.batch_processor'], 
                                        capture_output=True, text=True)
                if result.stdout.strip():
                    logging.info("Found existing batch processor running. Exiting.")
                    sys.exit(0)


            # Process each video
            for i, video_info in enumerate(videos, 1):
                logging.info(f"{'='*40}")
                logging.info(f"🎬 Processing video {i}/{len(videos)}")
                logging.info(f"{'='*40}")
                
                await self.process_single_video(video_info)
                
            
            logging.info("🏁 Playlist processing completed!")
            
        except Exception as e:
            logging.error(f"❌ Error processing playlist: {e}")

def main():
    # Initialize logging system
    setup_logging()
    
    logging.info("🎬 YouTube Video Processing Automation Started")
    logging.info("=" * 80)
    
    if sys.platform == "darwin":
        result = subprocess.run(['pgrep', '-f', '/youtube_video_processor/main.py'], 
                                capture_output=True, text=True)
        if result.stdout.strip():
            logging.info("Found existing cron job running. Exiting. \n \n \n \n")
            sys.exit(0)


    """Main entry point"""
    try:
        logging.info("🎬 YouTube Video Processing Automation Started")
        #logging.info("="*80)
        
        # Step 1: Ensure Ollama is running
        if not OllamaManager.is_ollama_running():
            if not OllamaManager.start_ollama():
                logging.error("❌ Failed to start Ollama. Exiting.")
                sys.exit(1)
        
        # Step 2: Initialize and run video processor
        processor = VideoProcessor()
        asyncio.run(processor.process_playlist())
        
        logging.info("✅ All processing completed successfully! \n \n \n \n")
        
    except KeyboardInterrupt:
        logging.info("⏹️ Processing interrupted by user")
        sys.exit(0)
    except Exception as e:
        logging.error(f"❌ Fatal error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
